"""Génération XLSX (openpyxl) et PDF (reportlab) à partir de lignes tabulaires."""
from __future__ import annotations

import io

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

WAGADU_BROWN = colors.HexColor("#6E3C13")
WAGADU_SAND = colors.HexColor("#F0E4C8")


def to_xlsx(*, title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.style = "Headline 3"
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(*, title: str, headers: list[str], rows: list[list]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=14 * mm, bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    header = Paragraph(f"<b>Wagadu Hub — {title}</b>", styles["Title"])

    data = [headers] + [["" if v is None else str(v) for v in r] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), WAGADU_BROWN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, WAGADU_SAND]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    doc.build([header, table])
    return buffer.getvalue()
